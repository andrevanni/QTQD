"""Endpoints do Excesso Crítico — assistente que processa o Excel de estoque por fabricante,
calcula o excesso por curva (A/B/C/D) e permite atribuir os totais a um lançamento existente."""
from datetime import datetime, timezone
from uuid import UUID

from fastapi import APIRouter, Body, Depends, Form, HTTPException, UploadFile

from backend.app.core.auth import get_current_tenant
from backend.app.db.client import get_supabase

router = APIRouter(prefix="/me/excesso-critico", tags=["excesso-critico"])

DEFAULT_LIMITES = {"limite_a": 120, "limite_b": 150, "limite_c": 150, "limite_d": 180}
CURVA_KEYS = {"A": "excesso_curva_a", "B": "excesso_curva_b", "C": "excesso_curva_c", "D": "excesso_curva_d"}
APLICAR_FIELDS = list(CURVA_KEYS.values()) + ["total_estoque_lancamentos"]


def _merge_aplicar_valores(valores: dict, payload: dict) -> dict:
    """Grava os campos aplicáveis (excesso por curva + total de lançamentos) em `valores`,
    preservando os demais. Lança ValueError se algum campo presente for não-numérico."""
    out = dict(valores)
    for field in APLICAR_FIELDS:
        v = payload.get(field)
        if v is not None:
            try:
                out[field] = float(v)
            except (TypeError, ValueError):
                raise ValueError(f"{field} deve ser numérico")
    return out


# ── Limites por curva (config do tenant) ──────────────────────────────────────

@router.get("/limites")
def obter_limites(tenant_id: UUID = Depends(get_current_tenant)) -> dict:
    res = get_supabase().table("tenant_excesso_config").select("*").eq("tenant_id", str(tenant_id)).limit(1).execute()
    if res.data:
        row = res.data[0]
        return {k: int(row.get(k, DEFAULT_LIMITES[k])) for k in DEFAULT_LIMITES}
    return dict(DEFAULT_LIMITES)


@router.put("/limites")
def salvar_limites(
    payload: dict = Body(...),
    tenant_id: UUID = Depends(get_current_tenant),
) -> dict:
    cleaned: dict = {}
    for k in DEFAULT_LIMITES:
        v = payload.get(k)
        try:
            n = int(v)
        except (TypeError, ValueError):
            n = DEFAULT_LIMITES[k]
        if n < 1 or n > 9999:
            raise HTTPException(status_code=400, detail=f"{k} deve estar entre 1 e 9999 dias")
        cleaned[k] = n

    sb = get_supabase()
    now = datetime.now(timezone.utc).isoformat()
    existing = sb.table("tenant_excesso_config").select("tenant_id").eq("tenant_id", str(tenant_id)).limit(1).execute()
    if existing.data:
        sb.table("tenant_excesso_config").update({**cleaned, "updated_at": now}).eq("tenant_id", str(tenant_id)).execute()
    else:
        sb.table("tenant_excesso_config").insert({"tenant_id": str(tenant_id), **cleaned, "updated_at": now}).execute()
    return cleaned


# ── Cálculo a partir do Excel ─────────────────────────────────────────────────

def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _norm_curva(v) -> str:
    return str(v or "").strip().upper()[:1]


@router.post("/calcular")
def calcular(
    file: UploadFile,
    limite_a: int = Form(120),
    limite_b: int = Form(150),
    limite_c: int = Form(150),
    limite_d: int = Form(180),
    tenant_id: UUID = Depends(get_current_tenant),
) -> dict:
    """Recebe o Excel `excesso_tabela_fabricante_*.xlsx` (colunas: Nome Completo, Linha, Curva,
    Filial, MediaF Un, Qtd Estoque, Estoque Valor) + limites por curva. Retorna totais agregados
    por curva, lista de produtos críticos e contagem."""
    import io
    from openpyxl import load_workbook

    limites = {"A": limite_a, "B": limite_b, "C": limite_c, "D": limite_d}
    for c, lim in limites.items():
        if not isinstance(lim, int) or lim < 1 or lim > 9999:
            raise HTTPException(status_code=400, detail=f"Limite da curva {c} deve estar entre 1 e 9999 dias")

    try:
        content = file.file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo inválido. Use o modelo Excel correto.")

    ws = wb.active

    # ── Lê cabeçalho e mapeia colunas pelo nome (tolerante a reordenação) ──
    header = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    def col(name_candidates: list[str]) -> int | None:
        for i, h in enumerate(header):
            if any(h.lower() == c.lower() for c in name_candidates):
                return i
        return None

    i_nome = col(["Nome Completo", "Nome"])
    i_linha = col(["Linha"])
    i_curva = col(["Curva"])
    i_media = col(["MediaF Un", "Media", "Media Un", "MediaF"])
    i_qtd = col(["Qtd Estoque", "Qtd", "Estoque Qtd"])
    i_valor = col(["Estoque Valor", "Valor", "Estoque R$"])

    if None in (i_nome, i_curva, i_media, i_qtd, i_valor):
        raise HTTPException(
            status_code=400,
            detail="Cabeçalho inválido. Esperado: Nome Completo, Linha, Curva, Filial, MediaF Un, Qtd Estoque, Estoque Valor.",
        )

    # ── Agrega por (nome, linha, curva) somando filiais ──
    agg: dict[tuple, dict] = {}
    total_linhas = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        nome = row[i_nome]
        if not nome:
            continue
        curva = _norm_curva(row[i_curva])
        if curva not in limites:
            continue
        linha = str(row[i_linha] or "").strip() if i_linha is not None else ""
        media = _to_float(row[i_media])
        qtd = _to_float(row[i_qtd])
        valor = _to_float(row[i_valor])

        key = (str(nome).strip(), linha, curva)
        if key in agg:
            a = agg[key]
            a["qtd"] += qtd
            a["media"] += media
            a["valor"] += valor
        else:
            agg[key] = {"nome": str(nome).strip(), "linha": linha, "curva": curva, "qtd": qtd, "media": media, "valor": valor}
        total_linhas += 1

    # ── Calcula excesso por produto ──
    totais_por_curva = {"A": 0.0, "B": 0.0, "C": 0.0, "D": 0.0}
    qtd_criticos_por_curva = {"A": 0, "B": 0, "C": 0, "D": 0}
    produtos_criticos: list[dict] = []
    valor_total_estoque = 0.0

    for item in agg.values():
        valor_total_estoque += item["valor"]
        qtd, media, valor, curva = item["qtd"], item["media"], item["valor"], item["curva"]
        if qtd <= 0:
            continue

        custo_un = valor / qtd if qtd > 0 else 0.0
        excesso_un = 0.0
        cobertura_dias: float | None = None

        if media == 0:
            # Sem venda e estoque > 1 → todo o estoque é excesso crítico
            if qtd > 1:
                excesso_un = qtd
                cobertura_dias = None  # infinita
        else:
            cobertura_dias = (qtd / media) * 30.0
            limite = limites[curva]
            if cobertura_dias > limite:
                ideal = (media * limite) / 30.0
                excesso_un = qtd - ideal

        if excesso_un > 0:
            excesso_valor = excesso_un * custo_un
            totais_por_curva[curva] += excesso_valor
            qtd_criticos_por_curva[curva] += 1
            produtos_criticos.append({
                "nome": item["nome"],
                "linha": item["linha"],
                "curva": curva,
                "qtd_estoque": round(qtd, 2),
                "media_un": round(media, 4),
                "cobertura_dias": round(cobertura_dias, 1) if cobertura_dias is not None else None,
                "excesso_un": round(excesso_un, 2),
                "custo_un": round(custo_un, 4),
                "excesso_valor": round(excesso_valor, 2),
            })

    produtos_criticos.sort(key=lambda p: p["excesso_valor"], reverse=True)

    total_excesso = sum(totais_por_curva.values())
    return {
        "limites": limites,
        "totais": {
            "excesso_curva_a": round(totais_por_curva["A"], 2),
            "excesso_curva_b": round(totais_por_curva["B"], 2),
            "excesso_curva_c": round(totais_por_curva["C"], 2),
            "excesso_curva_d": round(totais_por_curva["D"], 2),
            "total": round(total_excesso, 2),
        },
        "resumo": {
            "total_linhas_excel": total_linhas,
            "total_produtos_unicos": len(agg),
            "total_produtos_criticos": len(produtos_criticos),
            "qtd_criticos_por_curva": qtd_criticos_por_curva,
            "valor_total_estoque": round(valor_total_estoque, 2),
            "pct_excesso": round((total_excesso / valor_total_estoque) * 100, 2) if valor_total_estoque > 0 else 0.0,
        },
        "produtos": produtos_criticos[:100],  # top 100 para preview
    }


# ── Aplicar resultado a uma avaliação existente ──────────────────────────────

@router.post("/aplicar/{avaliacao_id}")
def aplicar(
    avaliacao_id: UUID,
    payload: dict = Body(...),
    tenant_id: UUID = Depends(get_current_tenant),
) -> dict:
    """Atualiza os campos excesso_curva_a/b/c/d no JSONB `valores` da avaliação especificada.
    Preserva todos os outros campos."""
    sb = get_supabase()
    res = (
        sb.table("avaliacoes_semanais")
        .select("id, valores, status, semana_referencia")
        .eq("id", str(avaliacao_id))
        .eq("tenant_id", str(tenant_id))
        .limit(1)
        .execute()
    )
    if not res.data:
        raise HTTPException(status_code=404, detail="Avaliação não encontrada.")

    valores = dict(res.data[0].get("valores") or {})
    try:
        valores = _merge_aplicar_valores(valores, payload)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    now = datetime.now(timezone.utc).isoformat()
    sb.table("avaliacoes_semanais").update({"valores": valores, "updated_at": now}).eq("id", str(avaliacao_id)).eq("tenant_id", str(tenant_id)).execute()

    return {
        "ok": True,
        "avaliacao_id": str(avaliacao_id),
        "semana_referencia": res.data[0]["semana_referencia"],
        "status": res.data[0]["status"],
        "valores_aplicados": {f: valores.get(f, 0) for f in APLICAR_FIELDS},
    }
