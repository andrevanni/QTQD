from datetime import datetime, timezone
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, UploadFile

from backend.app.core.auth import get_current_tenant
from backend.app.db.client import get_supabase
from backend.app.schemas.avaliacoes import (
    AvaliacaoCreateRequest,
    AvaliacaoResponse,
    AvaliacaoUpdateRequest,
    AvaliacaoValores,
)
from backend.app.services.calculos_qtqd import calcular_indicadores

router = APIRouter(prefix="/avaliacoes", tags=["avaliacoes"])

_COLS = "id, tenant_id, semana_referencia, status, observacoes, valores, created_at, updated_at"

# Campos persistidos no JSONB `valores` que NÃO são gerenciados pelo formulário
# nem pela importação Excel — apenas o assistente de Excesso Crítico (/aplicar) os
# escreve. Devem ser preservados em qualquer escrita que substitua `valores` inteiro
# (PATCH do formulário, import-excel), senão voltam ao default 0 e o valor aplicado
# é perdido silenciosamente.
APPLY_ONLY_VALORES_FIELDS = ("total_estoque_lancamentos",)


def _preserve_apply_only(new_valores: dict, old_valores: dict | None) -> dict:
    """Copia os campos apply-only do registro existente para o novo dict de valores,
    quando presentes — evita zerar valores gravados só pelo /aplicar."""
    if not old_valores:
        return new_valores
    for field in APPLY_ONLY_VALORES_FIELDS:
        if old_valores.get(field) is not None:
            new_valores[field] = old_valores[field]
    return new_valores

# Campos do template Excel: (chave_interna | None=seção, label, formato)
_EXCEL_CAMPOS = [
    (None, "QT — QUANTO TENHO", None),
    ("saldo_bancario",             "Saldo bancário",          "currency"),
    ("estoque_custo",              "Estoque (preço custo)",   "currency"),
    (None, "QT — Contas a receber (detalhado)", None),
    ("cartoes",                    "Cartões",                 "currency"),
    ("convenios",                  "Convênios",               "currency"),
    ("cheques",                    "Cheques",                 "currency"),
    ("trade_marketing",            "Trade marketing",         "currency"),
    ("outros_qt",                  "Outros QT",               "currency"),
    (None, "QD — Contas a pagar (detalhado)", None),
    ("fornecedores",               "Fornecedores",            "currency"),
    ("investimentos_assumidos",    "Investimentos assumidos", "currency"),
    ("outras_despesas_assumidas",  "Outras despesas assumidas","currency"),
    (None, "QD — Dívidas (detalhado)", None),
    ("financiamentos",             "Financiamentos",          "currency"),
    ("tributos_atrasados",         "Tributos atrasados",      "currency"),
    ("acoes_processos",            "Ações e processos",       "currency"),
    (None, "INFORMAÇÕES COMPLEMENTARES", None),
    ("faturamento_previsto_mes",   "Faturamento previsto no mês",                    "currency"),
    ("compras_mes",                "Compras no mês",                                 "currency"),
    ("entrada_mes",                "Entrada no mês",                                 "currency"),
    ("venda_cupom_mes",            "Venda cupom no mês",                             "currency"),
    ("venda_custo_mes",            "Venda custo no mês (CMV)",                       "currency"),
    ("lucro_liquido_mes",          "Lucro líquido no mês",                           "currency"),
    (None, "INDICADORES OPERACIONAIS", None),
    ("pmp",        "PMP — Prazo Médio de Pagamento (dias)",    "days"),
    ("pme_excel",  "PME — Cobertura Média (dias)",             "days"),
    ("pmv",        "PMV — Prazo Médio de Venda Total (dias)",  "days"),
    ("pmv_avista", "  PMV À Vista",                            "days"),
    ("pmv_30",     "  PMV 30 dias",                            "days"),
    ("pmv_60",     "  PMV 60 dias",                            "days"),
    ("pmv_90",     "  PMV 90 dias",                            "days"),
    ("pmv_120",    "  PMV 120 dias",                           "days"),
    ("pmv_outros", "  PMV Outros",                             "days"),
    ("cobertura_estoque_dia", "Cobertura de estoque (do dia)", "days"),
    ("indice_faltas", "Índice de Faltas (%)",                  "percent"),
    (None, "EXCESSO DE ESTOQUE", None),
    ("excesso_curva_a", "Excesso Curva A >90 dias",  "currency"),
    ("excesso_curva_b", "Excesso Curva B >120 dias", "currency"),
    ("excesso_curva_c", "Excesso Curva C >150 dias", "currency"),
    ("excesso_curva_d", "Excesso Curva D >180 dias", "currency"),
]


def _serialize(row: dict) -> AvaliacaoResponse:
    valores = AvaliacaoValores(**(row.get("valores") or {}))
    return AvaliacaoResponse(
        id=row["id"],
        tenant_id=row["tenant_id"],
        semana_referencia=row["semana_referencia"],
        status=row["status"],
        observacoes=row.get("observacoes"),
        created_at=row["created_at"],
        updated_at=row["updated_at"],
        valores=valores,
        indicadores=calcular_indicadores(valores),
    )


@router.get("/template-excel")
def template_excel(
    semana: str | None = None,
    tenant_id: UUID = Depends(get_current_tenant),
):
    """Gera e baixa um template Excel para lançamento semanal, opcionalmente pré-preenchido."""
    import io
    from datetime import date as date_type
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    sb = get_supabase()
    valores_db: dict = {}
    semana_ref = semana or ""

    if semana:
        res = (
            sb.table("avaliacoes_semanais")
            .select("semana_referencia,valores")
            .eq("tenant_id", str(tenant_id))
            .eq("semana_referencia", semana)
            .limit(1)
            .execute()
        )
        if res.data:
            valores_db = res.data[0].get("valores") or {}
            semana_ref = res.data[0]["semana_referencia"]

    data_display = ""
    if semana_ref:
        try:
            d = date_type.fromisoformat(semana_ref)
            data_display = d.strftime("%d/%m/%Y")
        except Exception:
            data_display = semana_ref

    wb = Workbook()
    ws = wb.active
    ws.title = "Lançamento QTQD"

    BLUE, BLUE_LT, DARK, SEC_BG, WHITE, GREY, ALT = (
        "2563EB", "DBEAFE", "1E3A5F", "EFF6FF", "FFFFFF", "999999", "F8FAFF"
    )

    def _font(bold=False, size=10, color="000000", italic=False):
        return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

    def _fill(c):
        return PatternFill("solid", fgColor=c)

    # Título
    ws.merge_cells("A1:C1")
    ws["A1"].value = "QTQD — Lançamento Semanal"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = _fill(DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Instrução
    ws.merge_cells("A2:C2")
    ws["A2"].value = "Preencha a coluna 'Valor' e importe de volta no portal QTQD."
    ws["A2"].font = _font(size=9, color=GREY, italic=True)
    ws["A2"].fill = _fill("F0F4FF")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Data da semana
    ws["A3"].value = "data_semana"
    ws["A3"].font = _font(size=8, color="AAAAAA")
    ws["B3"].value = "Data da semana (dd/mm/aaaa)"
    ws["B3"].font = _font(bold=True, color=DARK)
    ws["B3"].fill = _fill(BLUE_LT)
    ws["C3"].value = data_display
    ws["C3"].font = _font(bold=True, size=11, color=DARK)
    ws["C3"].fill = _fill(BLUE_LT)
    ws.row_dimensions[3].height = 20

    # Cabeçalhos das colunas
    for col, lbl in [(1, "Chave interna"), (2, "Campo"), (3, "Valor")]:
        c = ws.cell(row=4, column=col)
        c.value = lbl
        c.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        c.fill = _fill(BLUE)
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[4].height = 16

    cur = 5
    alt = 0
    for key, label, fmt in _EXCEL_CAMPOS:
        if key is None:
            ws.merge_cells(f"A{cur}:C{cur}")
            ws[f"A{cur}"].value = label
            ws[f"A{cur}"].font = _font(bold=True, size=9, color=BLUE)
            ws[f"A{cur}"].fill = _fill(SEC_BG)
            ws[f"A{cur}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[cur].height = 14
            cur += 1
            alt = 0
            continue

        bg = ALT if alt % 2 == 0 else WHITE
        alt += 1
        is_sub = label.startswith("  ")

        ws.cell(row=cur, column=1).value = key
        ws.cell(row=cur, column=1).font = _font(size=8, color="AAAAAA")
        ws.cell(row=cur, column=1).fill = _fill(bg)

        ws.cell(row=cur, column=2).value = label
        ws.cell(row=cur, column=2).font = _font(italic=is_sub, color="555555" if is_sub else "000000")
        ws.cell(row=cur, column=2).fill = _fill(bg)
        if is_sub:
            ws.cell(row=cur, column=2).alignment = Alignment(indent=2)

        raw = valores_db.get(key)
        if raw is not None and raw != 0:
            cell_val: float | None = round(float(raw) * 100, 4) if fmt == "percent" else float(raw)
        else:
            cell_val = None

        vc = ws.cell(row=cur, column=3)
        vc.value = cell_val
        vc.font = _font(bold=True)
        vc.fill = _fill(bg)
        vc.number_format = "#,##0.00" if fmt == "currency" else ("0.00" if fmt == "percent" else "0.##")
        ws.row_dimensions[cur].height = 18
        cur += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.freeze_panes = "C5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"qtqd_{semana or 'template'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("/import-excel", response_model=AvaliacaoResponse, status_code=201)
def import_excel(
    file: UploadFile,
    tenant_id: UUID = Depends(get_current_tenant),
):
    """Importa um arquivo Excel gerado pelo template QTQD e cria/atualiza o lançamento."""
    import io
    from openpyxl import load_workbook

    content = file.file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo inválido. Use o template gerado pelo portal QTQD.")

    ws = wb.active

    def _parse(v, is_pct: bool = False) -> float:
        """Converte valor do Excel (número ou string pt-BR) para float."""
        if v is None or v == "":
            return 0.0
        if isinstance(v, (int, float)):
            n = float(v)
            return round(n / 100, 6) if is_pct else n
        s = str(v).strip().replace("R$", "").replace(" ", "").replace("%", "")
        # pt-BR: "1.234,56" → "1234.56"
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        try:
            n = float(s)
            return round(n / 100, 6) if is_pct else n
        except ValueError:
            return 0.0

    semana_referencia: str | None = None
    valores: dict = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 3:
            continue
        key = row[0]
        value = row[2]
        if not key or key in ("Chave interna", "chave"):
            continue
        if key == "data_semana":
            if value:
                if isinstance(value, str) and "/" in value:
                    parts = value.strip().split("/")
                    if len(parts) == 3:
                        semana_referencia = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                elif hasattr(value, "strftime"):
                    semana_referencia = value.strftime("%Y-%m-%d")
            continue
        if value is not None and value != "":
            valores[key] = _parse(value, is_pct=(key == "indice_faltas"))

    if not semana_referencia:
        raise HTTPException(
            status_code=400,
            detail="Data da semana não encontrada. Verifique o campo 'Data da semana (dd/mm/aaaa)' no arquivo.",
        )

    sb = get_supabase()
    existing = (
        sb.table("avaliacoes_semanais")
        .select("id, valores")
        .eq("tenant_id", str(tenant_id))
        .eq("semana_referencia", semana_referencia)
        .limit(1)
        .execute()
    )

    av = AvaliacaoValores(**{k: v for k, v in valores.items() if hasattr(AvaliacaoValores, k) or k in AvaliacaoValores.model_fields})
    now = datetime.now(timezone.utc).isoformat()

    if existing.data:
        result = (
            sb.table("avaliacoes_semanais")
            .update({
                "valores": _preserve_apply_only(av.model_dump(), existing.data[0].get("valores")),
                "status": "rascunho",
                "updated_at": now,
            })
            .eq("id", existing.data[0]["id"])
            .eq("tenant_id", str(tenant_id))
            .execute()
        )
    else:
        result = (
            sb.table("avaliacoes_semanais")
            .insert({
                "tenant_id": str(tenant_id),
                "semana_referencia": semana_referencia,
                "status": "rascunho",
                "valores": av.model_dump(),
            })
            .execute()
        )

    if not result.data:
        raise HTTPException(status_code=500, detail="Falha ao salvar registro importado.")
    return _serialize(result.data[0])


@router.get("", response_model=list[AvaliacaoResponse])
def listar(tenant_id: UUID = Depends(get_current_tenant)) -> list[AvaliacaoResponse]:
    result = (
        get_supabase()
        .table("avaliacoes_semanais")
        .select(_COLS)
        .eq("tenant_id", str(tenant_id))
        .order("semana_referencia", desc=True)
        .execute()
    )
    return [_serialize(row) for row in result.data]


@router.get("/{avaliacao_id}", response_model=AvaliacaoResponse)
def obter(avaliacao_id: UUID, tenant_id: UUID = Depends(get_current_tenant)) -> AvaliacaoResponse:
    result = (
        get_supabase()
        .table("avaliacoes_semanais")
        .select(_COLS)
        .eq("id", str(avaliacao_id))
        .eq("tenant_id", str(tenant_id))
        .limit(1)
        .execute()
    )
    if not result.data:
        raise HTTPException(status_code=404, detail="Avaliacao nao encontrada.")
    return _serialize(result.data[0])


@router.post("", response_model=AvaliacaoResponse, status_code=201)
def criar(payload: AvaliacaoCreateRequest, tenant_id: UUID = Depends(get_current_tenant)) -> AvaliacaoResponse:
    if payload.tenant_id != tenant_id:
        raise HTTPException(status_code=403, detail="tenant_id do payload nao confere com o token.")
    valores = AvaliacaoValores(**payload.model_dump(exclude={"tenant_id", "semana_referencia", "status", "observacoes"}))
    data = {
        "tenant_id": str(tenant_id),
        "semana_referencia": str(payload.semana_referencia),
        "status": payload.status,
        "observacoes": payload.observacoes,
        "valores": valores.model_dump(),
    }
    result = get_supabase().table("avaliacoes_semanais").insert(data).execute()
    if not result.data:
        raise HTTPException(status_code=500, detail="Falha ao criar avaliacao.")
    return _serialize(result.data[0])


@router.patch("/{avaliacao_id}", response_model=AvaliacaoResponse)
def atualizar(
    avaliacao_id: UUID,
    payload: AvaliacaoUpdateRequest,
    tenant_id: UUID = Depends(get_current_tenant),
) -> AvaliacaoResponse:
    from backend.app.services.relatorio_service import enviar_relatorio_para_tenant

    sb = get_supabase()
    current = sb.table("avaliacoes_semanais").select(_COLS).eq("id", str(avaliacao_id)).eq("tenant_id", str(tenant_id)).limit(1).execute()
    if not current.data:
        raise HTTPException(status_code=404, detail="Avaliacao nao encontrada.")
    row = current.data[0]
    old_status = row["status"]
    new_status = payload.status or old_status
    next_valores = AvaliacaoValores(**(payload.valores.model_dump() if payload.valores else row.get("valores") or {}))
    next_valores_dict = next_valores.model_dump()
    if payload.valores:
        next_valores_dict = _preserve_apply_only(next_valores_dict, row.get("valores"))
    update_data = {
        "semana_referencia": str(payload.semana_referencia) if payload.semana_referencia else row["semana_referencia"],
        "status": new_status,
        "observacoes": payload.observacoes if payload.observacoes is not None else row.get("observacoes"),
        "valores": next_valores_dict,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    result = sb.table("avaliacoes_semanais").update(update_data).eq("id", str(avaliacao_id)).eq("tenant_id", str(tenant_id)).execute()

    # Se status transicionou de rascunho → fechada, disparar e-mail (igual ao /fechar)
    if old_status == "rascunho" and new_status == "fechada":
        cfg_res = sb.table("tenant_pdf_config").select("ativo").eq("tenant_id", str(tenant_id)).limit(1).execute()
        envio_ativo = cfg_res.data[0].get("ativo", True) if cfg_res.data else True
        if envio_ativo:
            try:
                enviar_relatorio_para_tenant(str(tenant_id), sb, avaliacao_id=str(avaliacao_id), origem="fechar")
            except Exception:
                pass  # Não bloqueia o save — erro já registrado no email_log

    return _serialize(result.data[0])


@router.post("/{avaliacao_id}/fechar", response_model=AvaliacaoResponse)
def fechar(avaliacao_id: UUID, tenant_id: UUID = Depends(get_current_tenant)) -> AvaliacaoResponse:
    from backend.app.services.relatorio_service import enviar_relatorio_para_tenant

    sb = get_supabase()
    result = (
        sb.table("avaliacoes_semanais")
        .update({"status": "fechada", "updated_at": datetime.now(timezone.utc).isoformat()})
        .eq("id", str(avaliacao_id))
        .eq("tenant_id", str(tenant_id))
        .execute()
    )
    if not result.data:
        raise HTTPException(status_code=404, detail="Avaliacao nao encontrada.")

    cfg_res = sb.table("tenant_pdf_config").select("ativo").eq("tenant_id", str(tenant_id)).limit(1).execute()
    envio_ativo = (cfg_res.data[0].get("ativo", True) if cfg_res.data else True)
    if envio_ativo:
        try:
            enviar_relatorio_para_tenant(str(tenant_id), sb, avaliacao_id=str(avaliacao_id), origem="fechar")
        except Exception:
            pass  # Não bloqueia o fechamento — erro já registrado no email_log

    return _serialize(result.data[0])


@router.post("/{avaliacao_id}/finalizar")
def finalizar(avaliacao_id: UUID, tenant_id: UUID = Depends(get_current_tenant)) -> dict:
    """Marca a semana como finalizada e dispara o e-mail automático para todos os usuários."""
    from backend.app.services.relatorio_service import enviar_relatorio_para_tenant

    sb = get_supabase()
    check = sb.table("avaliacoes_semanais").select("id,status").eq("id", str(avaliacao_id)).eq("tenant_id", str(tenant_id)).limit(1).execute()
    if not check.data:
        raise HTTPException(status_code=404, detail="Avaliacao nao encontrada.")

    sb.table("avaliacoes_semanais").update(
        {"status": "finalizado", "updated_at": datetime.now(timezone.utc).isoformat()}
    ).eq("id", str(avaliacao_id)).eq("tenant_id", str(tenant_id)).execute()

    final = sb.table("avaliacoes_semanais").select(_COLS).eq("id", str(avaliacao_id)).limit(1).execute()
    avaliacao = _serialize(final.data[0])

    try:
        destinatarios = enviar_relatorio_para_tenant(str(tenant_id), sb, avaliacao_id=str(avaliacao_id), origem="finalizar")
    except Exception:
        destinatarios = []

    return {
        "avaliacao": avaliacao.model_dump(),
        "enviado_para": destinatarios,
        "n_destinatarios": len(destinatarios),
    }


@router.post("/{avaliacao_id}/reenviar-relatorio")
def reenviar_relatorio(
    avaliacao_id: UUID,
    tenant_id: UUID = Depends(get_current_tenant),
    email_teste: str | None = None,
) -> dict:
    """Reenvia o e-mail de relatório sem alterar o status da avaliação.
    Passar email_teste=addr@x.com restringe o envio a esse endereço apenas.
    """
    from backend.app.services.relatorio_service import enviar_relatorio_para_tenant

    sb = get_supabase()
    if not sb.table("avaliacoes_semanais").select("id").eq("id", str(avaliacao_id)).eq("tenant_id", str(tenant_id)).limit(1).execute().data:
        raise HTTPException(status_code=404, detail="Avaliacao nao encontrada.")

    try:
        destinatarios = enviar_relatorio_para_tenant(str(tenant_id), sb, email_teste=email_teste, avaliacao_id=str(avaliacao_id), origem="reenviar")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Falha ao enviar relatorio: {e}")

    if not destinatarios:
        raise HTTPException(status_code=400, detail="Nenhum usuario ativo com e-mail cadastrado para este tenant.")

    return {"enviado_para": destinatarios, "n_destinatarios": len(destinatarios)}


@router.delete("/{avaliacao_id}", status_code=204)
def excluir(avaliacao_id: UUID, tenant_id: UUID = Depends(get_current_tenant)) -> None:
    result = (
        get_supabase()
        .table("avaliacoes_semanais")
        .delete()
        .eq("id", str(avaliacao_id))
        .eq("tenant_id", str(tenant_id))
        .execute()
    )
    if not result.data:
        raise HTTPException(status_code=404, detail="Avaliacao nao encontrada.")
